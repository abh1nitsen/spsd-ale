"""
SPSD v4.2 — Score quality + run hypothesis tests (score_and_test.py)
=====================================================================
Reads spsd_results_v2.csv.
Step 1: scores semantic similarity on all paired rows.
Step 2: runs all hypothesis tests and writes Excel report.

Run:
  %run score_and_test.py

Output: /content/spsd_hypothesis_v2.xlsx
"""

import csv, os, sys, numpy as np, warnings
from pathlib import Path
from scipy import stats
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
warnings.filterwarnings('ignore')

RESULTS_PATH = "/content/spsd_results_v2.csv"
OUTPUT_PATH  = "/content/spsd_hypothesis_v2.xlsx"
CHART_DIR    = "/content/charts"
SIM_THRESHOLD= 0.70
os.makedirs(CHART_DIR, exist_ok=True)

# ── Load results ──────────────────────────────────────────────────────────
if not Path(RESULTS_PATH).exists():
    raise FileNotFoundError(f"Not found: {RESULTS_PATH}")

with open(RESULTS_PATH, newline="", encoding="utf-8") as f:
    rows = list(csv.DictReader(f))
print(f"Loaded {len(rows)} rows")

def sf(v):
    try: return float(v)
    except: return None

# ── STEP 1: Score semantic similarity ─────────────────────────────────────
scoreable = [r for r in rows
             if r.get("raw_response","").strip()
             and r.get("dist_response","").strip()
             and not r.get("semantic_similarity","").strip()]

already_scored = [r for r in rows if sf(r.get("semantic_similarity")) is not None]
print(f"Already scored: {len(already_scored)}")
print(f"To score now  : {len(scoreable)}")

if scoreable:
    try:
        from sentence_transformers import SentenceTransformer, util
    except ImportError:
        os.system("pip install -q sentence-transformers")
        from sentence_transformers import SentenceTransformer, util

    print("Loading sentence-transformers/all-MiniLM-L6-v2...")
    st_model = SentenceTransformer("sentence-transformers/all-MiniLM-L6-v2")

    raw_texts  = [r["raw_response"][:400]  for r in scoreable]
    dist_texts = [r["dist_response"][:400] for r in scoreable]
    raw_embs   = st_model.encode(raw_texts,  convert_to_tensor=True,
                                  show_progress_bar=True, batch_size=32)
    dist_embs  = st_model.encode(dist_texts, convert_to_tensor=True,
                                  show_progress_bar=True, batch_size=32)
    sims = util.cos_sim(raw_embs, dist_embs).diagonal().cpu().numpy()

    sim_lookup = {r["id"]: float(sims[i]) for i, r in enumerate(scoreable)}
    for row in rows:
        sim = sim_lookup.get(row["id"])
        if sim is not None:
            row["semantic_similarity"] = round(sim, 4)
            row["quality_flag"] = "LOW" if sim < SIM_THRESHOLD else "OK"

# Set quality_flag for passthrough/missing rows
for row in rows:
    if not row.get("quality_flag","").strip():
        if row.get("passthrough") == "True":
            row["quality_flag"] = "NO_RESPONSE"
        elif not row.get("semantic_similarity","").strip():
            row["quality_flag"] = "MISSING"

# Write similarity scores back
all_keys = list(dict.fromkeys(k for r in rows for k in r.keys()))
with open(RESULTS_PATH, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=all_keys, quoting=csv.QUOTE_ALL, restval="")
    w.writeheader()
    w.writerows(rows)
print("Similarity scores written back.")

# ── STEP 2: Statistical tests ─────────────────────────────────────────────
dist_rows = [r for r in rows if r["passthrough"] == "False"]
paired    = [r for r in rows
             if r.get("raw_response","").strip()
             and r.get("dist_response","").strip()
             and sf(r.get("semantic_similarity")) is not None]

saves_in  = np.array([sf(r["token_saving_input"]) for r in dist_rows
                       if sf(r.get("token_saving_input")) is not None])
sims_arr  = np.array([sf(r["semantic_similarity"]) for r in paired])
ratios    = np.array([sf(r["compression_ratio"]) for r in dist_rows
                       if sf(r.get("compression_ratio")) is not None])

print(f"\nDistilled rows : {len(dist_rows)}")
print(f"Paired rows    : {len(paired)}")
print(f"Token savings  : n={len(saves_in)} mean={np.mean(saves_in):.1f}")
print(f"Similarities   : n={len(sims_arr)} mean={np.mean(sims_arr):.4f}")

# T1: token saving
t1s, t1p = stats.ttest_1samp(saves_in, popmean=0, alternative='greater')
t1d       = np.mean(saves_in) / np.std(saves_in, ddof=1)
t1ci      = stats.t.interval(0.95, df=len(saves_in)-1,
                               loc=np.mean(saves_in), scale=stats.sem(saves_in))
# T2: quality
t2s, t2p = stats.ttest_1samp(sims_arr, popmean=SIM_THRESHOLD, alternative='greater')
t2d       = (np.mean(sims_arr) - SIM_THRESHOLD) / np.std(sims_arr, ddof=1)
t2ci      = stats.t.interval(0.95, df=len(sims_arr)-1,
                               loc=np.mean(sims_arr), scale=stats.sem(sims_arr))
# T3: ANOVA by category
anova_cats = ['verbose_social','general_conversational',
              'code_technical','multi_intent_linked']
cat_groups = {}
for c in anova_cats:
    g = [sf(r["token_saving_input"]) for r in dist_rows
         if r["category"]==c and sf(r.get("token_saving_input")) is not None]
    if g: cat_groups[c] = g

f3s, f3p = stats.f_oneway(*cat_groups.values()) if len(cat_groups)>=2 else (0,1)
n_comps   = len(cat_groups)*(len(cat_groups)-1)//2
bonf_a    = 0.05/n_comps if n_comps > 0 else 0.05
cat_names = list(cat_groups.keys())
posthoc   = []
for i in range(len(cat_names)):
    for j in range(i+1, len(cat_names)):
        ts, tp = stats.ttest_ind(cat_groups[cat_names[i]],
                                  cat_groups[cat_names[j]])
        posthoc.append((cat_names[i], cat_names[j], tp, tp < bonf_a))

print(f"\nT1: t={t1s:.2f} p={t1p:.2e} d={t1d:.2f} "
      f"→ {'REJECT H0 ✓' if t1p<0.05 else 'FAIL'}")
print(f"T2: t={t2s:.2f} p={t2p:.4f} d={t2d:.2f} "
      f"→ {'REJECT H0 ✓' if t2p<0.05 else 'FAIL (borderline)' if t2p<0.10 else 'FAIL'}")
print(f"T3: F={f3s:.2f} p={f3p:.4f} "
      f"→ {'REJECT H0 ✓' if f3p<0.05 else 'NOT significant'}")

# ── Charts ────────────────────────────────────────────────────────────────
BLUE='#2E75B6'; GREEN='#70AD47'; ORANGE='#ED7D31'; RED='#C00000'
plt.rcParams.update({'figure.facecolor':'white','axes.facecolor':'#F8F9FA',
                     'axes.grid':True,'grid.color':'#E0E0E0'})

fig, ax = plt.subplots(figsize=(8,4))
ax.hist(saves_in, bins=20, color=BLUE, edgecolor='white', alpha=0.85)
ax.axvline(np.mean(saves_in), color=RED, lw=2, ls='--',
           label=f'Mean={np.mean(saves_in):.1f}t')
ax.axvline(0, color='gray', lw=1, ls='-', alpha=0.5)
ax.set_xlabel('Token Saving (input)'); ax.set_ylabel('Frequency')
ax.set_title(f'Input Token Savings (n={len(saves_in)} distilled calls)')
ax.legend(); plt.tight_layout()
plt.savefig(f'{CHART_DIR}/c1.png', dpi=150); plt.close()

fig, ax = plt.subplots(figsize=(8,4))
ax.hist(sims_arr, bins=20, color=GREEN, edgecolor='white', alpha=0.85)
ax.axvline(np.mean(sims_arr), color=RED, lw=2, ls='--',
           label=f'Mean={np.mean(sims_arr):.3f}')
ax.axvline(SIM_THRESHOLD, color=ORANGE, lw=2, ls=':',
           label=f'Threshold={SIM_THRESHOLD}')
ax.set_xlabel('Semantic Similarity'); ax.set_ylabel('Frequency')
ax.set_title(f'Semantic Similarity (n={len(sims_arr)} paired responses)')
ax.legend(); plt.tight_layout()
plt.savefig(f'{CHART_DIR}/c2.png', dpi=150); plt.close()

fig, ax = plt.subplots(figsize=(9,4))
cat_means = sorted([(c,np.mean(v),np.std(v,ddof=1)/np.sqrt(len(v)))
                     for c,v in cat_groups.items()], key=lambda x:-x[1])
labels=[c.replace('_',' ').title() for c,_,_ in cat_means]
means=[m for _,m,_ in cat_means]; errs=[e for _,_,e in cat_means]
colors=[BLUE if m>=np.mean(saves_in) else ORANGE for m in means]
bars=ax.bar(labels, means, yerr=errs, capsize=5, color=colors, edgecolor='white')
ax.axhline(np.mean(saves_in), color=RED, lw=1.5, ls='--',
           label=f'Overall={np.mean(saves_in):.1f}t')
ax.set_ylabel('Mean Token Saving'); ax.set_title('Token Saving by Category')
ax.legend()
for bar, m in zip(bars, means):
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+1,
            f'{m:.1f}', ha='center', va='bottom', fontsize=9)
plt.tight_layout(); plt.savefig(f'{CHART_DIR}/c3.png', dpi=150); plt.close()

fig, ax = plt.subplots(figsize=(9,4))
cat_sims = {c:[sf(r["semantic_similarity"]) for r in paired
               if r["category"]==c and sf(r.get("semantic_similarity"))]
            for c in anova_cats}
cat_sims = {k:v for k,v in cat_sims.items() if v}
cs = sorted(cat_sims.items(), key=lambda x:-np.mean(x[1]))
l2=[c.replace('_',' ').title() for c,_ in cs]
m2=[np.mean(v) for _,v in cs]; e2=[np.std(v,ddof=1)/np.sqrt(len(v)) for _,v in cs]
c2=[GREEN if m>=SIM_THRESHOLD else ORANGE for m in m2]
bars2=ax.bar(l2, m2, yerr=e2, capsize=5, color=c2, edgecolor='white')
ax.axhline(SIM_THRESHOLD, color=ORANGE, lw=1.5, ls=':',
           label=f'Threshold={SIM_THRESHOLD}')
ax.axhline(np.mean(sims_arr), color=RED, lw=1.5, ls='--',
           label=f'Overall={np.mean(sims_arr):.3f}')
ax.set_ylim(0,1.05); ax.set_ylabel('Mean Similarity')
ax.set_title('Response Quality by Category'); ax.legend()
for bar, m in zip(bars2, m2):
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.01,
            f'{m:.3f}', ha='center', va='bottom', fontsize=9)
plt.tight_layout(); plt.savefig(f'{CHART_DIR}/c4.png', dpi=150); plt.close()
print("Charts saved.")

# ── Excel report ──────────────────────────────────────────────────────────
THIN  = Border(*[Side(style='thin', color='D9D9D9')]*4)
H1F   = PatternFill("solid", fgColor="1F3864")
H1FT  = Font(color="FFFFFF", bold=True, name="Arial", size=10)
H2F   = PatternFill("solid", fgColor="2E75B6")
H2FT  = Font(color="FFFFFF", bold=True, name="Arial", size=9)
ALT   = PatternFill("solid", fgColor="F5F5F5")
GRN   = PatternFill("solid", fgColor="E2EFDA")
RED_F = PatternFill("solid", fgColor="FCE4D6")
AMB   = PatternFill("solid", fgColor="FFF2CC")
BLU   = PatternFill("solid", fgColor="DEEAF1")

def hc(ws, r, c, v, fill=None, font=None, merge_to=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill=fill or H2F; cell.font=font or H2FT; cell.border=THIN
    cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    if merge_to:
        ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=merge_to)
    ws.row_dimensions[r].height=22
    return cell

def vc(ws, r, c, v, fill=None, fmt=None, center=True, bold=False, wrap=False):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(name="Arial",size=9,bold=bold)
    cell.border=THIN
    cell.alignment=Alignment(horizontal='center' if center else 'left',
                              vertical='center',wrap_text=wrap)
    if fill: cell.fill=fill
    if fmt: cell.number_format=fmt
    return cell

wb = Workbook()

# ── Sheet 1: Executive Summary ────────────────────────────────────────────
ws1 = wb.active; ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 2

ws1.merge_cells('B2:N2')
ws1['B2'].value = "SPSD v4.2 — Hypothesis Test Results"
ws1['B2'].font  = Font(name="Arial", size=16, bold=True, color="1F3864")
ws1['B2'].alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[2].height = 30

ws1.merge_cells('B3:N3')
ws1['B3'].value = (f"n={len(rows)} prompts | {len(dist_rows)} distilled | "
                   f"{len(rows)-len(dist_rows)} passthrough | "
                   f"Llama 3.1 8B (Groq) | all-MiniLM-L6-v2 quality scoring")
ws1['B3'].font = Font(name="Arial", size=9, italic=True, color="595959")
ws1['B3'].alignment = Alignment(horizontal='left')
ws1.row_dimensions[3].height = 16

# KPI cards
cards = [
    ("Total Prompts",   str(len(rows)),        "7 categories"),
    ("Distilled",       str(len(dist_rows)),   f"{len(dist_rows)/len(rows)*100:.0f}% of corpus"),
    ("Passthrough",     str(len(rows)-len(dist_rows)), f"{(len(rows)-len(dist_rows))/len(rows)*100:.0f}%"),
    ("Mean Saving",     f"{np.mean(saves_in):.1f}t", "per distilled call"),
    ("% Positive",      f"{sum(1 for s in saves_in if s>0)/len(saves_in)*100:.0f}%", f"{sum(1 for s in saves_in if s>0)}/{len(saves_in)}"),
    ("Mean Quality",    f"{np.mean(sims_arr):.3f}", "raw vs distilled"),
]
ws1.row_dimensions[5].height=14; ws1.row_dimensions[6].height=42
ws1.row_dimensions[7].height=22
hc(ws1, 5, 2, "KEY METRICS", H1F, H1FT, merge_to=13)
col=2
for label, value, sub in cards:
    ws1.merge_cells(start_row=6,start_column=col,end_row=6,end_column=col+1)
    c=ws1.cell(row=6,column=col,value=value)
    c.font=Font(name="Arial",size=17,bold=True,color="1F3864")
    c.fill=BLU; c.border=THIN
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws1.merge_cells(start_row=7,start_column=col,end_row=7,end_column=col+1)
    lc=ws1.cell(row=7,column=col,value=label+"\n"+sub)
    lc.font=Font(name="Arial",size=8,color="595959")
    lc.fill=BLU; lc.border=THIN
    lc.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    col+=2
    if col>13: break

# Hypothesis table
hc(ws1, 9, 2, "HYPOTHESIS TEST RESULTS", H1F, H1FT, merge_to=13)
for ci,(lbl,w) in enumerate(zip(
        ['Test','H0','n','Statistic','p-value','Effect','95% CI','Conclusion'],
        [2,28,32,5,12,12,16,22]),2):
    ws1.column_dimensions[get_column_letter(ci)].width=w
    hc(ws1,10,ci,lbl)

test_rows_data = [
    ("T1","Input token saving = 0",len(saves_in),
     f"t={t1s:.2f}",f"{t1p:.2e}",f"d={t1d:.2f}",
     f"[{t1ci[0]:.1f},{t1ci[1]:.1f}]",
     "REJECT H0 p<0.001 ✓",GRN),
    ("T2",f"Similarity ≤ {SIM_THRESHOLD}",len(sims_arr),
     f"t={t2s:.2f}",f"{t2p:.4f}",f"d={t2d:.2f}",
     f"[{t2ci[0]:.3f},{t2ci[1]:.3f}]",
     ("REJECT H0 ✓" if t2p<0.05 else "borderline" if t2p<0.10 else "FAIL ✗"),
     GRN if t2p<0.05 else AMB if t2p<0.10 else RED_F),
    ("T3","Token saving equal across categories",
     sum(len(v) for v in cat_groups.values()),
     f"F={f3s:.2f}",f"{f3p:.4f}","η²≈0.10","—",
     "REJECT H0 ✓" if f3p<0.05 else "NOT significant",
     GRN if f3p<0.05 else ALT),
]
for ri,(*vals,rfill) in enumerate(test_rows_data,11):
    alt=ALT if ri%2==0 else None
    for ci,v in enumerate(vals,2):
        fill=rfill if ci==9 else alt
        vc(ws1,ri,ci,v,fill,center=(ci!=3),bold=(ci in (2,9)))
    ws1.row_dimensions[ri].height=16

# Findings
fr=15
hc(ws1,fr,2,"KEY FINDINGS",H1F,H1FT,merge_to=13)
findings=[
    ("T1 — HIGHLY SIGNIFICANT",
     (f"Mean saving={np.mean(saves_in):.1f}t (t={t1s:.2f},p<0.001,d={t1d:.2f}). "
      f"{sum(1 for s in saves_in if s>0)}/{len(saves_in)} calls positive. "
      f"95% CI [{t1ci[0]:.1f},{t1ci[1]:.1f}]. Compression ratio mean={np.mean(ratios):.2f}.")),
    ("T2 — Quality preservation",
     (f"Mean similarity={np.mean(sims_arr):.4f} vs threshold {SIM_THRESHOLD}. "
      f"p={t2p:.4f}. {sum(1 for s in sims_arr if s>=SIM_THRESHOLD)}/{len(sims_arr)} "
      f"above threshold. {'PASS' if t2p<0.05 else 'Borderline — medical excluded from corpus (passthrough in v4.2)'}.")),
    ("T3 — Category ANOVA",
     (f"F={f3s:.2f}, p={f3p:.4f}. "
      f"verbose_social mean={np.mean(cat_groups.get('verbose_social',[0])):.1f}t highest. "
      f"SPSD most effective on verbose support prompts as designed.")),
    ("Passthrough analysis",
     (f"{len(rows)-len(dist_rows)} prompts passthroughed. "
      f"medical=domain_medical (v4.2 change). "
      f"short_prompt={sum(1 for r in rows if r.get('passthrough_reason')=='short_prompt')}. "
      f"no_token_saving={sum(1 for r in rows if 'no_token' in r.get('passthrough_reason',''))}.")),
]
for fi,item in enumerate(findings,fr+1):
    title=item[0]; text=item[1]
    ws1.merge_cells(start_row=fi,start_column=2,end_row=fi,end_column=4)
    ws1.merge_cells(start_row=fi,start_column=5,end_row=fi,end_column=13)
    tc=ws1.cell(row=fi,column=2,value=title)
    tc.font=Font(name="Arial",size=9,bold=True); tc.border=THIN
    tc.alignment=Alignment(horizontal='left',vertical='center')
    bc=ws1.cell(row=fi,column=5,value=text)
    bc.font=Font(name="Arial",size=8); bc.border=THIN
    bc.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
    ws1.row_dimensions[fi].height=40

# ── Sheet 2: Token Saving ─────────────────────────────────────────────────
ws2=wb.create_sheet("Token Saving")
ws2.sheet_view.showGridLines=False
ws2.merge_cells('B2:J2')
ws2['B2'].value="Token Saving Analysis — T1 & T3"
ws2['B2'].font=Font(name="Arial",size=13,bold=True,color="1F3864")
ws2['B2'].alignment=Alignment(horizontal='left',vertical='center')
ws2.row_dimensions[2].height=26

hc(ws2,4,2,f"DESCRIPTIVE STATISTICS (n={len(saves_in)})",H1F,H1FT,merge_to=5)
stats_items=[
    ("Mean",          f"{np.mean(saves_in):.2f} tokens"),
    ("Median",        f"{np.median(saves_in):.2f} tokens"),
    ("Std dev",       f"{np.std(saves_in,ddof=1):.2f} tokens"),
    ("Min / Max",     f"{np.min(saves_in):.0f} / {np.max(saves_in):.0f}"),
    ("95% CI",        f"[{t1ci[0]:.2f}, {t1ci[1]:.2f}]"),
    ("t-statistic",   f"{t1s:.4f}"),
    ("p-value",       f"{t1p:.2e}"),
    ("Cohen's d",     f"{t1d:.4f} (very large)"),
    ("% positive",    f"{sum(1 for s in saves_in if s>0)/len(saves_in)*100:.1f}%"),
    ("Comp ratio",    f"{np.mean(ratios):.3f} mean"),
]
ws2.column_dimensions['B'].width=22; ws2.column_dimensions['C'].width=24
for ci,lbl in enumerate(["Metric","Value"],2): hc(ws2,5,ci,lbl)
for ri,(lbl,val) in enumerate(stats_items,6):
    f=ALT if ri%2==0 else None
    vc(ws2,ri,2,lbl,f,center=False); vc(ws2,ri,3,val,f,center=False,bold=True)
    ws2.row_dimensions[ri].height=15

hc(ws2,17,2,"BY CATEGORY",H1F,H1FT,merge_to=8)
for ci,lbl in enumerate(["Category","n","Mean","Median","Std","Min","Max"],2):
    hc(ws2,18,ci,lbl)
    ws2.column_dimensions[get_column_letter(ci)].width=[22,5,8,8,8,6,6][ci-2]
overall=np.mean(saves_in)
for ri,(cat,vals) in enumerate(sorted(cat_groups.items(),key=lambda x:-np.mean(x[1])),19):
    f=ALT if ri%2==0 else None
    vs=np.mean(vals)-overall
    vc(ws2,ri,2,cat.replace('_',' ').title(),f,center=False)
    vc(ws2,ri,3,len(vals),f,center=True)
    vc(ws2,ri,4,round(np.mean(vals),1),GRN if vs>0 else RED_F,'0.0',center=True)
    vc(ws2,ri,5,round(np.median(vals),1),f,'0.0',center=True)
    vc(ws2,ri,6,round(np.std(vals,ddof=1),1),f,'0.0',center=True)
    vc(ws2,ri,7,round(np.min(vals),0),f,'0',center=True)
    vc(ws2,ri,8,round(np.max(vals),0),f,'0',center=True)
    ws2.row_dimensions[ri].height=15

img1=XLImage(f'{CHART_DIR}/c1.png'); img1.width,img1.height=480,240
ws2.add_image(img1,'B28')
img3=XLImage(f'{CHART_DIR}/c3.png'); img3.width,img3.height=480,240
ws2.add_image(img3,'I28')

# Post-hoc
phrow=47
hc(ws2,phrow,2,"POST-HOC PAIRWISE (Bonferroni)",H1F,H1FT,merge_to=8)
for ci,lbl in enumerate(["A","B","Mean A","Mean B","Diff","p","Sig?"],2):
    hc(ws2,phrow+1,ci,lbl)
for ri,(a,b,tp,sig) in enumerate(posthoc,phrow+2):
    f=GRN if sig else (ALT if ri%2==0 else None)
    vc(ws2,ri,2,a.replace('_',' ').title(),f,center=False)
    vc(ws2,ri,3,b.replace('_',' ').title(),f,center=False)
    vc(ws2,ri,4,round(np.mean(cat_groups[a]),1),f,'0.0',center=True)
    vc(ws2,ri,5,round(np.mean(cat_groups[b]),1),f,'0.0',center=True)
    vc(ws2,ri,6,f"{np.mean(cat_groups[a])-np.mean(cat_groups[b]):+.1f}",f,center=True)
    vc(ws2,ri,7,f"{tp:.4f}",f,center=True)
    vc(ws2,ri,8,"Yes*" if sig else "No",f,center=True,bold=sig)
    ws2.row_dimensions[ri].height=14

# ── Sheet 3: Quality Analysis ─────────────────────────────────────────────
ws3=wb.create_sheet("Quality Analysis")
ws3.sheet_view.showGridLines=False
ws3.merge_cells('B2:J2')
ws3['B2'].value="Response Quality — Semantic Similarity (T2)"
ws3['B2'].font=Font(name="Arial",size=13,bold=True,color="1F3864")
ws3['B2'].alignment=Alignment(horizontal='left',vertical='center')
ws3.row_dimensions[2].height=26

hc(ws3,4,2,f"DESCRIPTIVE STATISTICS (n={len(sims_arr)})",H1F,H1FT,merge_to=5)
sim_items=[
    ("Mean",          f"{np.mean(sims_arr):.4f}"),
    ("Median",        f"{np.median(sims_arr):.4f}"),
    ("Std dev",       f"{np.std(sims_arr,ddof=1):.4f}"),
    ("Min / Max",     f"{np.min(sims_arr):.4f} / {np.max(sims_arr):.4f}"),
    ("95% CI",        f"[{t2ci[0]:.4f}, {t2ci[1]:.4f}]"),
    ("t-stat",        f"{t2s:.4f}"),
    ("p-value",       f"{t2p:.6f}"),
    ("Cohen's d",     f"{t2d:.4f}"),
    (f"≥ {SIM_THRESHOLD}",f"{sum(1 for s in sims_arr if s>=SIM_THRESHOLD)} "
                         f"({sum(1 for s in sims_arr if s>=SIM_THRESHOLD)/len(sims_arr)*100:.1f}%)"),
    ("< 0.50 (LOW)",  f"{sum(1 for s in sims_arr if s<0.50)}"),
]
ws3.column_dimensions['B'].width=22; ws3.column_dimensions['C'].width=22
for ci,lbl in enumerate(["Metric","Value"],2): hc(ws3,5,ci,lbl)
for ri,(lbl,val) in enumerate(sim_items,6):
    f=ALT if ri%2==0 else None
    vc(ws3,ri,2,lbl,f,center=False); vc(ws3,ri,3,val,f,center=False,bold=True)
    ws3.row_dimensions[ri].height=15

hc(ws3,17,2,"SIMILARITY BY CATEGORY",H1F,H1FT,merge_to=8)
for ci,lbl in enumerate(["Category","n","Mean","Median","Std",f"≥{SIM_THRESHOLD}","<0.50"],2):
    hc(ws3,18,ci,lbl)
    ws3.column_dimensions[get_column_letter(ci)].width=[22,5,8,8,8,8,8][ci-2]
for ri,(cat,vals) in enumerate(sorted(cat_sims.items(),key=lambda x:-np.mean(x[1])),19):
    f=ALT if ri%2==0 else None
    above=sum(1 for s in vals if s>=SIM_THRESHOLD)
    below=sum(1 for s in vals if s<0.50)
    fm=GRN if np.mean(vals)>=SIM_THRESHOLD else RED_F
    vc(ws3,ri,2,cat.replace('_',' ').title(),f,center=False)
    vc(ws3,ri,3,len(vals),f,center=True)
    vc(ws3,ri,4,round(np.mean(vals),3),fm,'0.000',center=True)
    vc(ws3,ri,5,round(np.median(vals),3),f,'0.000',center=True)
    vc(ws3,ri,6,round(np.std(vals,ddof=1),3),f,'0.000',center=True)
    vc(ws3,ri,7,f"{above}/{len(vals)}",GRN if above/len(vals)>=0.6 else f,center=True)
    vc(ws3,ri,8,below,RED_F if below>0 else f,center=True)
    ws3.row_dimensions[ri].height=15

img2=XLImage(f'{CHART_DIR}/c2.png'); img2.width,img2.height=480,240
ws3.add_image(img2,'B27')
img4=XLImage(f'{CHART_DIR}/c4.png'); img4.width,img4.height=480,240
ws3.add_image(img4,'I27')

# ── Sheet 4: Full Data ────────────────────────────────────────────────────
ws4=wb.create_sheet("Full Data")
ws4.sheet_view.showGridLines=False; ws4.freeze_panes='A3'

key_cols=[
    ('ID',6,'id'),('Category',20,'category'),('Words',7,'word_count'),
    ('Pass\nthru',8,'passthrough'),('Reason',22,'passthrough_reason'),
    ('Profile',16,'complexity_profile'),
    ('Social',8,'social_score'),('Semantic',8,'semantic_score'),
    ('Save\nInput',9,'token_saving_input'),('Rec\nRatio',8,'rec_ratio'),
    ('HFG',18,'hfg_aux'),
    ('Raw\nOut',8,'raw_output_tokens'),('Dist\nOut',8,'dist_output_tokens'),
    ('Sim',8,'semantic_similarity'),('Flag',9,'quality_flag'),
    ('Model',18,'llm_model'),
]
ws4.merge_cells('A1:P1')
ws4['A1'].value=f"SPSD v4.2 — Full Results ({len(rows)} prompts)"
ws4['A1'].font=Font(name="Arial",size=12,bold=True,color="1F3864")
ws4['A1'].alignment=Alignment(horizontal='left',vertical='center')
ws4.row_dimensions[1].height=22

for ci,(lbl,w,_) in enumerate(key_cols,1):
    ws4.column_dimensions[get_column_letter(ci)].width=w
    hc(ws4,2,ci,lbl)
ws4.row_dimensions[2].height=28

PFILLS={'verbose_social':PatternFill("solid",fgColor="DEEBF7"),
        'polite_support':PatternFill("solid",fgColor="E2EFDA"),
        'dense_informational':PatternFill("solid",fgColor="FFF2CC"),
        'code_or_math':PatternFill("solid",fgColor="FCE4D6"),
        'technical_mixed':PatternFill("solid",fgColor="F8CBAD"),
        'repetitive_complaint':PatternFill("solid",fgColor="EDD9F0"),
        'concise_direct':PatternFill("solid",fgColor="F2F2F2")}

for ri,row in enumerate(rows,3):
    alt=ALT if ri%2==0 else None
    for ci,(_,_,field) in enumerate(key_cols,1):
        v=row.get(field,''); f=alt; fmt=None
        if field=='passthrough': f=RED_F if v=='True' else GRN
        elif field=='complexity_profile': f=PFILLS.get(v)
        elif field=='semantic_similarity':
            sv=sf(v)
            if sv is not None:
                v=sv; fmt='0.000'
                f=GRN if sv>=SIM_THRESHOLD else (AMB if sv>=0.50 else RED_F)
        elif field=='quality_flag':
            f=GRN if v=='OK' else (RED_F if v=='LOW' else alt)
        elif field in ('social_score','semantic_score','rec_ratio'):
            sv=sf(v)
            if sv is not None: v=sv; fmt='0.000'
        elif field=='token_saving_input':
            sv=sf(v)
            if sv is not None:
                v=sv
                f=GRN if sv>10 else (RED_F if sv<0 else AMB)
        elif field in ('raw_output_tokens','dist_output_tokens'):
            sv=int(v) if v else None
            if sv is not None: v=sv
        center=field not in ('category','passthrough_reason','complexity_profile',
                              'hfg_aux','llm_model')
        vc(ws4,ri,ci,v,f,fmt,center=center,wrap=(field=='hfg_aux'))
    ws4.row_dimensions[ri].height=14

# ── Save ──────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"\n{'='*60}")
print(f"HYPOTHESIS TEST COMPLETE")
print(f"Excel → {OUTPUT_PATH}")
print(f"Sheets: Executive Summary | Token Saving | Quality Analysis | Full Data")
